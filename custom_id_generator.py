"""
    This script is used to read an Excel file, add a "Custom ID" column, and then generate the 
    Custom ID based on the values of the "Author" column and "Title" column.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nameparser import HumanName


FILE_PATH = Path("Dev Testing.xlsx")


# Read the excel file into the primary DataFrame
df = pd.read_excel(
    FILE_PATH, sheet_name="Sheet1", usecols=("Author", "Title", "Custom ID")
)


def extract_first_letters(title):
    """
    This function takes a title string and extracts the first letter of each word,
    while preserving numbers. 'The' is excluded if its the first word.
    Non-alphanumeric characters and spaces are ignored.

    Parameters:
    - title (str): The input title from which to extract first letters.

    Returns:
    str: A string containing the first letters of each word and/or numbers all together.

    Example:
    >>> extract_first_letters("The Lord of Rings")
    'LotR'
    >>> extract_first_letters("101 Dalimations")
    '101D'
    """

    words = re.findall(r"\b[\w\d]+\b", title)

    if words and words[0].lower() == "the":
        words = words[1:]

    return "".join(word[0] if word.isalpha() else word for word in words)


def process_author_initials(row):
    """
    Process the 'Author' column in the given DataFrame row to extract and format author initials.

    This checks for the presence of the '&' character in the 'Author' column and processes
    the author information accordingly. It assigns formatted initials to the 'initials' column.

    Parameters:
    - row (pd.Series): A single row of the DataFrame containing 'Author' information.

    Returns:
    pd.Series: The row with updated data in column 'initials'.

     Example:
    >>> process_author_initials("Austin, Jane")
    'JA'
    >>> process_author_initials("Sanderson, Brandon & Patterson, Janci")
    'BS&JP '
    """

    # Check if the condition is met for this row
    authors = row["Author"].split("&")

    # Process each author and their initials
    initials = []
    for author in authors:
        author = author.strip()
        initials.append(HumanName(author).initials().replace(".", "").replace(" ", ""))

    # Combine initials when multiple authors are present
    row["initials"] = "&".join(initials)

    return row


# Apply the function to each row
result_df = df.apply(process_author_initials, axis=1)


# Count number of books by author and iterate up by one.
def count_books(input_df):
    """
    Adds a "Count" column and assigns a unique two-digit count for each author's occurrences.

    Args:
        input_df (pd.DataFrame): DataFrame containing an "Author" column.

    Returns:
        pd.DataFrame: The modified DataFrame with the "Count" column added and values added.
    """
    count_df = (
        input_df.copy()
    )  # Make a copy to avoid modifying the input DataFrame directly
    count_df["Count"] = (
        result_df.groupby("Author").cumcount().add(1).astype(str).str.zfill(2)
    )
    return count_df


result_df = count_books(result_df)


# Create new Custom ID
def create_custom_id(row):
    """
    Creates a new "Custom ID" column by combining the "initials", "Title", and "Count" columns.

    Args:
        row (pd.Series): A single row of the df with "initials", "Title", and "Count" columns.

    Returns:
        pd.Series: The row with the new "Custom ID" column added.
    """
    row["Custom ID"] = (
        row["initials"] + "_" + extract_first_letters(row["Title"]) + "_" + row["Count"]
    )
    return row


result_df = result_df.apply(create_custom_id, axis=1)


workbook = load_workbook(FILE_PATH)
sheet = workbook["Sheet1"]


# Writes the "Custom ID" values from the DataFrame to original spreadsheet, starting in cell C2.
for index, r in enumerate(result_df["Custom ID"], start=2):
    sheet.cell(row=index, column=3, value=r)

# Save the file with the Custom ID data populated.
workbook.save(FILE_PATH)
